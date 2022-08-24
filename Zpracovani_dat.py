# -*- coding: utf-8 -*-
"""
Created on Thu Aug 11 08:29:03 2022

@author: BilKu

Upraví data získaná od ČHMÚ do vhodnějšího formátu.
"""

import openpyxl
from openpyxl import Workbook
import datetime

class IKO:
    def IKO_by_formula(Pollutants, Relative_values):
        # Vypočítává IKO ze zadaných hodnot dle vzorce v Index_kvality_ovzdusi.pdf
        
        prekracujici = 0
        prekracujici_pocet = 0
        v_norme = 0
        v_norme_pocet = 0
        
        for i in range(len(Pollutants)):
            if Pollutants[i] > Relative_values[i]:
                prekracujici = prekracujici + Pollutants[i]/Relative_values[i]
                prekracujici_pocet += 1
            else:
                v_norme = v_norme + Pollutants[i]/Relative_values[i]
                v_norme_pocet += 1
                
            if prekracujici_pocet == 0:
                index = v_norme/v_norme_pocet
            else:
                if v_norme_pocet == 0:
                    index = prekracujici/prekracujici_pocet
                else:
                    index = prekracujici/prekracujici_pocet + v_norme/v_norme_pocet
        return index    

    def IKO_by_CHMU(Pollutants, Relative_values):
        # Vypočítává IKO ze zadaných hodnot dle vzorce, který používá ČHMÚ
        
        index = 0
        for i in range(len(Pollutants)):
            index = index + Pollutants[i]/Relative_values[i]
            
        return index

    def IKO_main():
        # Doplní data_2018.xlsx sloupcem s IKO
        
        Relative_values = [350, 200, 90, 120]
    
        data_2018 = openpyxl.load_workbook("data_2018.xlsx")
        sheet = data_2018.active
        sheet["Q1"] = "IKO"
    
        pollutants = openpyxl.load_workbook("Skodliviny_2018.xlsx")
        list_of_quantities = ["SO2", "NO2", "PM10", "O3"]
    
        transposed_pollutant = openpyxl.Workbook()
        for i in list_of_quantities:
        
            current_sheet = pollutants.get_sheet_by_name(i)
            transposed_sheet = transposed_pollutant.create_sheet(i)
    
            for row in range(1, current_sheet.max_row + 1):
                for col in range(1, current_sheet.max_column + 1):
                    transposed_sheet.cell(col, row).value = current_sheet.cell(row, col).value
    
        date = datetime.date(2018, 1, 1)
        current_row = 2
    
        while date.year != 2019:
        
            row = date.day +2
            column = chr(date.month + 66)
            pos = str(column) + str(row)
            Pollutants = []
    
            for i in list_of_quantities:
                quantity = transposed_pollutant.get_sheet_by_name(i)
                pollutant = quantity[pos].value
                if pollutant == None:
                    pollutant = quantity[ str(column) + "33" ].value
                Pollutants.append(pollutant)
            
            index = IKO.IKO_by_CHMU(Pollutants, Relative_values)
            sheet["Q" + str(current_row)] = index
        
            current_row += 1
            date = date + datetime.timedelta(days = 1)
    
        data_2018.save("data_2018.xlsx")
        
class weather:
    def add_weather():
        # Vytvoří soubor data_2018.xlsx s vhodněji uspořádanými daty o počasí
        
        data_2018 = Workbook()
        sheet = data_2018.active
        
        sheet["A1"] = "date"
        row = 2
        date = datetime.date(2018, 1, 1)
        
        while date.year != 2019: 
        
            sheet["A" + str(row)] = date
            date = date + datetime.timedelta(days = 1)
            row += 1
        
        weather = openpyxl.load_workbook("CHMU_pocasi_all.xlsx")
        
        list_of_quantities = ["teplota průměrná", "rychlost větru", "tlak vzduchu", "vlhkost vzduchu", "úhrn srážek"]
        column = 66
        
        for i in list_of_quantities:
            
            sheet[str(chr(column)) + "1"] = str(i)
            sheet[str(chr(column + 1)) + "1"] = str(i) + "^2"
            sheet[str(chr(column + 2)) + "1"] = str(i) + "^3"
            
            quantity = weather.get_sheet_by_name(i)
            start_row = 5
            a = quantity["A5"]
        
            while a.value != "2018":
                start_row += 1
                a = quantity["A" + str(start_row)]
        
            b = 2
            for r in quantity.iter_rows(start_row, start_row + 11, 3, 34, True):
                for cell in r:
                    if cell != None:
                        sheet[str(chr(column)) + str(b)] = cell
                        sheet[str(chr(column + 1)) + str(b)] = cell**2
                        sheet[str(chr(column + 2)) + str(b)] = cell**3
                        b += 1
            column += 3
        data_2018.save("data_2018.xlsx")

def main():
    weather.add_weather()
    IKO.IKO_main()

main()
